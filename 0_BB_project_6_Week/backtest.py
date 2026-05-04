import sys
import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import sqlite3
import talib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import FormulaRule

# Импорт конфигурации
try:
    import config
except ImportError:
    print("Ошибка: Не найден файл config.py. Убедитесь, что он находится в той же папке.")
    sys.exit(1)

# Настройки
DB_PATH = config.DB_PATH
TICKERS = config.TICKERS
BB_PERIOD = config.BB_PERIOD
BB_STD_DEV = config.BB_STD_DEV
TAKE_PROFIT_PERCENT = config.TAKE_PROFIT_PERCENT
STOP_LOSS_PERCENT = getattr(config, 'STOP_LOSS_PERCENT', 0)  # Опционально
COMMISSION_PERCENT = getattr(config, 'COMMISSION_PERCENT', 0.05)  # По умолчанию 0.05%

class BacktestEngine:
    def __init__(self, start_date_str):
        self.start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        self.conn = sqlite3.connect(DB_PATH)
        self.trades_log = []
        self.portfolio = {ticker: {'qty': 0, 'avg_price': 0.0, 'invested': 0.0} for ticker in TICKERS}
        self.cash = 1000000.0  # Стартовый капитал (условный)
        self.initial_cash = self.cash
        
    def get_trading_data(self, ticker, start_date):
        """Загружает данные свечи из БД, начиная с первой доступной даты после start_date"""
        query = """
            SELECT date, open, high, low, close, volume 
            FROM candles 
            WHERE ticker = ? AND date >= ? 
            ORDER BY date ASC
        """
        df = pd.read_sql_query(query, self.conn, params=(ticker, start_date.strftime('%Y-%m-%d')))
        
        if df.empty:
            return None
            
        # Проверка на наличие данных именно в дату старта. Если нет - ищем ближайший будущий день
        if df.iloc[0]['date'] != start_date.strftime('%Y-%m-%d'):
            first_date = datetime.strptime(df.iloc[0]['date'], '%Y-%m-%d')
            print(f"⚠️ Для {ticker} нет данных на {start_date.date()}. Начало бэктеста перенесено на {first_date.date()} ({first_date.strftime('%A')})")
            
        df['date'] = pd.to_datetime(df['date'])
        return df

    def calculate_indicators(self, df):
        """Расчет индикаторов Боллинджера"""
        close = df['close'].values
        upper, middle, lower = talib.BBANDS(close, timeperiod=BB_PERIOD, nbdevup=BB_STD_DEV, nbdevdn=BB_STD_DEV, matype=0)
        df['bb_upper'] = upper
        df['bb_middle'] = middle
        df['bb_lower'] = lower
        return df

    def execute_backtest(self):
        print(f"\n🚀 Запуск бэктеста с {self.start_date.date()}...")
        print(f"💰 Стартовый капитал: {self.initial_cash:,.2f} RUB")
        print(f"📊 Акции: {', '.join(TICKERS)}\n")

        # Собираем все данные заранее, чтобы синхронизировать даты
        all_data = {}
        min_date = None
        
        for ticker in TICKERS:
            df = self.get_trading_data(ticker, self.start_date)
            if df is not None:
                df = self.calculate_indicators(df)
                all_data[ticker] = df
                if min_date is None or df['date'].min() < min_date:
                    min_date = df['date'].min()
        
        if not all_data:
            print("❌ Нет данных ни по одной акции в указанном периоде.")
            return

        # Объединяем даты для общего цикла
        all_dates = pd.concat([df['date'] for df in all_data.values()]).unique()
        all_dates.sort()
        
        # Основной цикл по дням
        for current_date in all_dates:
            date_str = current_date.strftime('%Y-%m-%d')
            
            # 1. Сначала проверяем условия продажи (Take Profit / Stop Loss) по текущим ценам
            for ticker in TICKERS:
                if ticker not in all_data: continue
                df = all_data[ticker]
                row = df[df['date'] == current_date]
                if row.empty: continue
                
                current_price = row.iloc[0]['close']
                pos = self.portfolio[ticker]
                
                if pos['qty'] > 0:
                    # Проверка Take Profit
                    profit_pct = (current_price - pos['avg_price']) / pos['avg_price'] * 100
                    
                    should_sell = False
                    reason = ""
                    
                    if profit_pct >= TAKE_PROFIT_PERCENT:
                        should_sell = True
                        reason = "TP"
                    elif STOP_LOSS_PERCENT > 0 and profit_pct <= -STOP_LOSS_PERCENT:
                        should_sell = True
                        reason = "SL"
                    
                    if should_sell:
                        # Продажа всей позиции
                        revenue = pos['qty'] * current_price
                        commission = revenue * (COMMISSION_PERCENT / 100)
                        net_revenue = revenue - commission
                        
                        profit = net_revenue - (pos['qty'] * pos['avg_price'])
                        
                        self.log_trade(ticker, date_str, "SELL", pos['qty'], current_price, profit, reason)
                        
                        self.cash += net_revenue
                        pos['qty'] = 0
                        pos['avg_price'] = 0.0
                        pos['invested'] = 0.0

            # 2. Генерация сигналов на покупку (только если нет позиции или есть место для усреднения)
            for ticker in TICKERS:
                if ticker not in all_data: continue
                df = all_data[ticker]
                row = df[df['date'] == current_date]
                if row.empty: continue
                
                close = row.iloc[0]['close']
                bb_lower = row.iloc[0]['bb_lower']
                bb_middle = row.iloc[0]['bb_middle']
                
                pos = self.portfolio[ticker]
                
                # Логика входа (упрощенная версия вашей стратегии)
                # Сигнал на покупку: цена пробила нижнюю границу
                if close < bb_lower:
                    # Если позиции нет -> КУПИ
                    if pos['qty'] == 0:
                        self.open_position(ticker, date_str, close, "BUY_INIT")
                    # Если позиция есть, но цена еще ниже -> ДОКУПИ (усреднение)
                    # Ограничим усреднение, например, до 3 уровней или пока есть кэш
                    elif pos['qty'] > 0 and pos['qty'] < 1000: 
                         self.open_position(ticker, date_str, close, "BUY_ADD")
                
                # Сигнал на выход по тренду (цена выше средней) - если хотим фиксировать не только по ТП
                # Но в вашей логике основная фиксация по % от средней. Оставим как есть.

        self.generate_report()

    def open_position(self, ticker, date_str, price, signal_type):
        """Открытие или усреднение позиции"""
        # Размер покупки: фиксируем сумму или кол-во? Возьмем фиксированную долю от свободного кэша или фикс кол-во
        # Для бэктеста возьмем фиксированное количество лотов, например 10, если хватает денег
        qty_to_buy = 10 
        cost = qty_to_buy * price
        commission = cost * (COMMISSION_PERCENT / 100)
        total_cost = cost + commission
        
        if self.cash < total_cost:
            # Не хватает денег на полную лотность, пробуем купить сколько можем
            qty_to_buy = int((self.cash * 0.95) / price) # оставляем запас на комиссию
            if qty_to_buy <= 0: return
            cost = qty_to_buy * price
            commission = cost * (COMMISSION_PERCENT / 100)
            total_cost = cost + commission

        pos = self.portfolio[ticker]
        
        # Расчет новой средней цены
        old_value = pos['qty'] * pos['avg_price']
        new_value = cost
        new_qty = pos['qty'] + qty_to_buy
        new_avg = (old_value + new_value) / new_qty if new_qty > 0 else 0
        
        pos['qty'] = new_qty
        pos['avg_price'] = new_avg
        pos['invested'] += total_cost
        
        self.cash -= total_cost
        
        self.log_trade(ticker, date_str, "BUY", qty_to_buy, price, 0, signal_type)

    def log_trade(self, ticker, date, action, qty, price, profit, reason):
        self.trades_log.append({
            'Date': date,
            'Ticker': ticker,
            'Action': action,
            'Qty': qty,
            'Price': round(price, 2),
            'Profit': round(profit, 2),
            'Reason': reason
        })

    def generate_report(self):
        if not self.trades_log:
            print("❌ Сделки не совершены. Проверьте параметры стратегии или период.")
            return

        df_trades = pd.DataFrame(self.trades_log)
        
        # --- Расчет метрик ---
        
        # Фильтруем только закрытые сделки (SELL) для расчета прибыли
        sells = df_trades[df_trades['Action'] == 'SELL']
        total_profit = sells['Profit'].sum()
        total_trades = len(sells)
        
        win_trades = sells[sells['Profit'] > 0]
        loss_trades = sells[sells['Profit'] <= 0]
        
        win_rate = (len(win_trades) / total_trades * 100) if total_trades > 0 else 0
        
        gross_profit = win_trades['Profit'].sum() if not win_trades.empty else 0
        gross_loss = abs(loss_trades['Profit'].sum()) if not loss_trades.empty else 0
        
        profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else float('inf')
        
        # Итоговый капитал
        # Считаем стоимость открытых позиций по последней цене
        final_cash = self.cash
        for ticker, pos in self.portfolio.items():
            if pos['qty'] > 0 and ticker in config.TICKERS:
                # Берем последнюю цену из БД
                last_price_df = pd.read_sql_query("SELECT close FROM candles WHERE ticker=? ORDER BY date DESC LIMIT 1", self.conn, params=(ticker,))
                if not last_price_df.empty:
                    final_cash += pos['qty'] * last_price_df.iloc[0]['close']
        
        total_return = ((final_cash - self.initial_cash) / self.initial_cash) * 100
        
        # Максимальная просадка (упрощенно по эквити кривой)
        # Строим кривую капитала по дням
        equity_curve = [self.initial_cash]
        current_eq = self.initial_cash
        # Сортируем сделки по дате и считаем сальдо
        # Это упрощенный расчет, так как у нас нет ежедневного баланса, только точки сделок
        # Для точности нужно было бы прогонять баланс каждый день, но возьмем пиковую прибыль и текущий минимум
        
        # Доп метрики
        avg_win = win_trades['Profit'].mean() if not win_trades.empty else 0
        avg_loss = loss_trades['Profit'].mean() if not loss_trades.empty else 0
        avg_trade_duration = "N/A" # Требует сложной логики парования купил-продал по тикеру
        
        # Статистика по акциям
        stats_by_ticker = sells.groupby('Ticker').agg(
            Total_Profit=('Profit', 'sum'),
            Trade_Count=('Profit', 'count'),
            Win_Rate=('Profit', lambda x: (x > 0).sum() / len(x) * 100)
        ).round(2)
        
        # --- Создание Excel отчета ---
        wb = Workbook()
        wb.remove(wb.active)
        
        # Лист 1: Сводка
        ws_summary = wb.create_sheet("Сводка")
        summary_data = [
            ["Метрика", "Значение"],
            ["Стартовая дата", self.start_date.strftime('%Y-%m-%d')],
            ["Дата отчета", datetime.now().strftime('%Y-%m-%d %H:%M')],
            ["Начальный капитал", f"{self.initial_cash:,.2f}"],
            ["Конечный капитал (с позициями)", f"{final_cash:,.2f}"],
            ["Общая доходность (%)", f"{total_return:.2f}%"],
            ["Количество сделок (Sell)", total_trades],
            ["Win Rate (%)", f"{win_rate:.2f}%"],
            ["Profit Factor", f"{profit_factor:.2f}" if profit_factor != float('inf') else "Inf"],
            ["Валовая прибыль", f"{gross_profit:,.2f}"],
            ["Валовый убыток", f"{gross_loss:,.2f}"],
            ["Средняя прибыль на сделку", f"{(total_profit/total_trades if total_trades else 0):,.2f}"],
            ["Макс. прибыль в сделке", f"{sells['Profit'].max():,.2f}" if not sells.empty else "0"],
            ["Макс. убыток в сделке", f"{sells['Profit'].min():,.2f}" if not sells.empty else "0"],
        ]
        
        for r_idx, row in enumerate(summary_data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                if c_idx == 2 and isinstance(value, str) and '%' in value:
                    cell.number_format = '0.00%'
        
        # Лист 2: По акциям
        ws_tickers = wb.create_sheet("По акциям")
        ws_tickers.append(["Тикер", "Доходность", "Сделок", "Win Rate %"])
        for ticker in stats_by_ticker.index:
            row = stats_by_ticker.loc[ticker]
            ws_tickers.append([
                ticker, 
                row['Total_Profit'], 
                row['Trade_Count'], 
                row['Win_Rate']
            ])
        
        # Лист 3: Все сделки
        ws_details = wb.create_sheet("Все сделки")
        for r_idx, row in enumerate(dataframe_to_rows(df_trades, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_details.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                
                # Раскраска прибыли
                if c_idx == 6 and r_idx > 1: # Column Profit
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")

        # Автоширина колонок
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50)

        filename = f"backtest_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        print(f"✅ Отчет сохранен в файл: {filename}")
        
        # Вывод в консоль
        print("\n" + "="*40)
        print("📊 КРАТКИЙ ОТЧЕТ ПО БЭКТЕСТУ")
        print("="*40)
        print(f"Общая доходность: {total_return:.2f}%")
        print(f"Всего сделок: {total_trades}")
        print(f"Win Rate: {win_rate:.2f}%")
        print(f"Profit Factor: {profit_factor:.2f}")
        print("\nДоходность по акциям:")
        print(stats_by_ticker.to_string())
        print("="*40)

def main():
    if len(sys.argv) > 1:
        start_date = sys.argv[1]
    else:
        start_date = input("Введите дату начала бэктеста (ГГГГ-ММ-ДД): ")
    
    try:
        # Проверка формата даты
        datetime.strptime(start_date, "%Y-%m-%d")
        engine = BacktestEngine(start_date)
        engine.execute_backtest()
    except ValueError:
        print("❌ Ошибка формата даты. Используйте ГГГГ-ММ-ДД (например, 2025-01-01)")
    except Exception as e:
        print(f"❌ Произошла ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
